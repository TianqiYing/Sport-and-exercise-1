import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor
import matplotlib.pyplot as plt
import scipy.stats as stats
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from xgboost import XGBRegressor

pd.set_option("display.max_columns", None)
pd.set_option("display.width", None)
pd.set_option("display.max_colwidth", None)

SEASON = "2024-25"
PREV_SEASON = "2023-24"
TARGET_GW = 38
BUDGET = 830
FORM_WINDOW = 3
EWMA_ALPHA = 0.1
ROLLING_WINDOW = 5

constraints = {"GK": 1, "DEF": 4, "MID": 4, "FWD": 2}
MAX_PER_TEAM = 3
MC_ITER = 20000

GRAPH_START_GW = 1
GRAPH_END_GW = 38
GWs = GRAPH_END_GW - GRAPH_START_GW + 1
df = pd.read_csv("active_perfect_understat_enhanced.csv", encoding="latin1", low_memory=False)

df = df[
    (df["season_x"] == SEASON) |
    (df["season_x"] == PREV_SEASON)
].copy()

df = df.sort_values(["player", "season_x", "gameweek"])

def map_position(pos):
    if pos == "GK":
        return "GK"
    elif pos in ["DC", "DL", "DR", "DMC", "DML", "DMR"]:
        return "DEF"
    elif pos in ["MC", "ML", "MR", "AMC", "AML", "AMR"]:
        return "MID"
    elif pos in ["FW", "FWL", "FWR"]:
        return "FWD"
    elif pos == "Sub":
        return np.nan
    return np.nan

df["position_mapped"] = df["position"].apply(map_position)

df["position_mapped"] = df.groupby("player")["position_mapped"].transform(
    lambda x: x.fillna(x.mode().iloc[0] if not x.mode().empty else "MID")
)

df = df[df["minutes_x"] > 0]

# FEATURES (NO LEAKAGE)
df["points_ewma"] = df.groupby("player")["total_points"].transform(
    lambda x: x.shift(1).ewm(alpha=EWMA_ALPHA).mean()
)
df["minutes_ewma"] = df.groupby("player")["minutes_x"].transform(
    lambda x: x.shift(1).ewm(alpha=EWMA_ALPHA).mean()
)
df["creativity_ewma"] = df.groupby("player")["creativity"].transform(
    lambda x: x.shift(1).ewm(alpha=EWMA_ALPHA).mean()
)
df["threat_ewma"] = df.groupby("player")["threat"].transform(
    lambda x: x.shift(1).ewm(alpha=EWMA_ALPHA).mean()
)

df["recent_points"] = df.groupby("player")["total_points"].transform(
    lambda x: x.shift(1).rolling(FORM_WINDOW, min_periods=1).mean()
)
df["recent_goals"] = df.groupby("player")["goals_scored"].transform(
    lambda x: x.shift(1).rolling(FORM_WINDOW, min_periods=1).mean()
)
df["recent_assists"] = df.groupby("player")["assists_x"].transform(
    lambda x: x.shift(1).rolling(FORM_WINDOW, min_periods=1).mean()
)

# STD
df["std_rolling"] = df.groupby("player")["total_points"].transform(
    lambda x: x.shift(1).rolling(FORM_WINDOW, min_periods=2).std()
)
df["std_expanding"] = df.groupby("player")["total_points"].transform(
    lambda x: x.shift(1).expanding().std()
)

USE_EXPANDING_STD = True
df["std_points"] = df["std_expanding"] if USE_EXPANDING_STD else df["std_rolling"]

df["std_points"] = df.groupby("player")["std_points"].transform(
    lambda x: x.fillna(x.expanding().mean())
)

df["std_points"] = df["std_points"].fillna(df["std_points"].median()).replace(0, 0.4)

#features = [ "points_ewma", "minutes_ewma","creativity_ewma","threat_ewma","recent_points","recent_goals","recent_assists",]

#model = RandomForestRegressor(n_estimators=200, random_state=42)

df = df.loc[:, ~df.columns.duplicated()]
df = df.drop(columns=[
    "league_id", "player_id", "position_id", "minutes_y",
    "own_goals_y", "game_id", "season_id",
    "game", "venue",
    "ea_index", "loaned_in", "loaned_out",
    "kickoff_time_formatted",
    "id",
], errors="ignore")
df = df.sort_values(["player", "date"]).reset_index(drop=True)
df = df[df["season_x"] != "2025-26"].reset_index(drop=True)
# Compute league position per team per gameweek

df = df.drop(columns=[c for c in df.columns if c.endswith("_drop")])
print(f"Duplicate columns : {df.columns[df.columns.duplicated()].tolist()}")
print(f"Duplicate indices : {df.index.duplicated().sum()}")

# =============================================================================
# 2. OPPONENT STRENGTH FEATURES
# =============================================================================
team_level = (
    df.groupby(["team", "date", "opponent", "season_x"], as_index=False)
    .agg(
        goals_scored_team = ("goals_scored", "sum"),
        xg_team           = ("xg",           "sum"),
    )
)

opp_conceded = (
    team_level
    .rename(columns={
        "team"             : "opponent",
        "opponent"         : "team",
        "goals_scored_team": "opp_goals_conceded",
        "xg_team"          : "opp_xg_conceded",
    })
    [["team", "date", "opp_goals_conceded", "opp_xg_conceded"]]
)

team_level = team_level.merge(opp_conceded, on=["team", "date"], how="left")
team_level = team_level.sort_values(["team", "date"]).reset_index(drop=True)

for w in [3, 5]:
    team_level[f"opp_goals_conceded_form_{w}"] = (
        team_level.groupby("team")["opp_goals_conceded"]
        .transform(lambda x: x.shift(1).rolling(w, min_periods=1).mean())
    )
    team_level[f"opp_xg_conceded_form_{w}"] = (
        team_level.groupby("team")["opp_xg_conceded"]
        .transform(lambda x: x.shift(1).rolling(w, min_periods=1).mean())
    )

assert team_level.duplicated(subset=["team", "date"]).sum() == 0, \
    "Duplicate (team, date) rows in team_level"

opp_form_cols = [
    "opp_goals_conceded_form_3", "opp_xg_conceded_form_3",
    "opp_goals_conceded_form_5", "opp_xg_conceded_form_5",
]

df = df.merge(
    team_level[["team", "date"] + opp_form_cols],
    left_on  = ["opponent", "date"],
    right_on = ["team", "date"],
    how      = "left",
    suffixes = ("", "_drop"),
)

df = df.drop(columns=[c for c in df.columns if c.endswith("_drop")])
df = df.loc[:, ~df.columns.duplicated()].reset_index(drop=True)

print(f"After opp merge — duplicate indices: {df.index.duplicated().sum()}")
# Compute league position per team per gameweek
# Points = 3 for win, 1 for draw, 0 for loss
df["match_points"] = np.where(
    df["goals_scored"] > df["goals_conceded"], 3,
    np.where(df["goals_scored"] == df["goals_conceded"], 1, 0)
)

league_table = (
    df.groupby(["team", "season_x", "gameweek"], as_index=False)
    .agg(
        cumulative_pts = ("match_points", "sum"),
        goal_diff      = ("goals_scored",  "sum"),
    )
)

# Cumulative sum up to but not including current GW (shift to avoid leakage)
league_table = league_table.sort_values(["season_x", "team", "gameweek"])
league_table["cumulative_pts"] = (
    league_table.groupby(["season_x", "team"])["cumulative_pts"]
    .transform(lambda x: x.shift(1).cumsum().fillna(0))
)

# Rank within season and gameweek
league_table["league_position"] = (
    league_table.groupby(["season_x", "gameweek"])["cumulative_pts"]
    .rank(ascending=False, method="min")
)

# Merge own team position
df = df.merge(
    league_table[["team", "season_x", "gameweek", "league_position"]],
    on=["team", "season_x", "gameweek"], how="left"
).rename(columns={"league_position": "own_league_position"})

# Merge opponent position
df = df.merge(
    league_table[["team", "season_x", "gameweek", "league_position"]],
    left_on=["opponent", "season_x", "gameweek"],
    right_on=["team", "season_x", "gameweek"],
    how="left", suffixes=("", "_drop")
).rename(columns={"league_position": "opp_league_position"})
df = df.drop(columns=[c for c in df.columns if c.endswith("_drop")])
# =============================================================================
# 3. SEASON-LEVEL FEATURES
# =============================================================================
season_order = sorted(df["season_x"].unique())
season_rank  = {s: i for i, s in enumerate(season_order)}
df["season_rank"] = df["season_x"].map(season_rank)

season_summary = (
    df.groupby(["player", "season_x", "season_rank"], as_index=False)
    .agg(
        prev_season_pts     = ("total_points", "sum"),
        prev_season_goals   = ("goals_scored",  "sum"),
        prev_season_assists = ("assists_x",      "sum"),
        prev_season_minutes = ("minutes_x",      "sum"),
        prev_season_xg      = ("xg",             "mean"),
        prev_season_xa      = ("xa",             "mean"),
        prev_season_games   = ("played",          "sum"),
    )
)

season_summary["season_rank"] += 1
season_summary = season_summary.drop_duplicates(
    subset=["player", "season_rank"], keep="first"
)

df = df.merge(
    season_summary[[
        "player", "season_rank",
        "prev_season_pts", "prev_season_goals", "prev_season_assists",
        "prev_season_minutes", "prev_season_xg", "prev_season_xa",
        "prev_season_games",
    ]],
    on=["player", "season_rank"],
    how="left"
)

df["prev_season_complete"] = (df["prev_season_games"] >= 25).astype(float)
df = df.loc[:, ~df.columns.duplicated()].reset_index(drop=True)

print(f"After season merge — duplicate indices: {df.index.duplicated().sum()}")

# =============================================================================
# 4. TEAM-LEVEL INTER-PLAYER FEATURES
# Captures the collective quality of teammates around each player.
# All rolling windows are shift(1) to avoid leaking current GW data.
# =============================================================================
print("\nBuilding team-level inter-player features...")

# Aggregate per team per game — one row per (team, date)
team_context = (
    df.groupby(["team", "date", "season_x"], as_index=False)
    .agg(
        # Attacking context — how dangerous is this team collectively?
        team_xg_sum_3          = ("form_xg_3",          "sum"),
        team_xg_sum_5          = ("form_xg_5",          "sum"),
        team_shots_sum_3       = ("form_shots_3",        "sum"),
        team_shots_sum_5       = ("form_shots_5",        "sum"),
        team_goals_sum_3       = ("form_goals_3",        "sum"),
        team_goals_sum_5       = ("form_goals_5",        "sum"),

        # Creativity context — how well does this team create chances?
        team_xa_sum_3          = ("form_xa_3",           "sum"),
        team_xa_sum_5          = ("form_xa_5",           "sum"),
        team_creativity_sum_3  = ("form_creativity_3",   "sum"),
        team_creativity_sum_5  = ("form_creativity_5",   "sum"),
        team_key_passes_sum_3  = ("form_key_passes_y_3", "sum"),
        team_key_passes_sum_5  = ("form_key_passes_y_5", "sum"),

        # Influence context — general contribution of the squad
        team_influence_sum_3   = ("form_influence_3",    "sum"),
        team_influence_sum_5   = ("form_influence_5",    "sum"),
        team_bps_sum_3         = ("form_bps_3",          "sum"),
        team_bps_sum_5         = ("form_bps_5",          "sum"),

        # In-form teammates — how many teammates are scoring/assisting?
        teammates_scoring_3    = ("form_goals_3",   lambda x: (x > 0.3).sum()),
        teammates_scoring_5    = ("form_goals_5",   lambda x: (x > 0.3).sum()),
        teammates_assisting_3  = ("form_assists_3", lambda x: (x > 0.2).sum()),
        teammates_assisting_5  = ("form_assists_5", lambda x: (x > 0.2).sum()),

        # Minutes context — how many teammates are getting regular minutes?
        teammates_playing_3    = ("form_minutes_3", lambda x: (x > 45).sum()),
        teammates_playing_5    = ("form_minutes_5", lambda x: (x > 45).sum()),

        # Squad size tracked (useful normalisation denominator)
        squad_size             = ("player", "count"),
    )
)

# Sort and apply rolling shift to avoid leaking current GW into features
team_context = team_context.sort_values(["team", "date"]).reset_index(drop=True)

# Normalise by squad size to make features comparable across teams
# (a team with 15 players tracked vs 10 shouldn't look artificially better)
norm_cols = [
    "team_xg_sum_3", "team_xg_sum_5",
    "team_shots_sum_3", "team_shots_sum_5",
    "team_goals_sum_3", "team_goals_sum_5",
    "team_xa_sum_3", "team_xa_sum_5",
    "team_creativity_sum_3", "team_creativity_sum_5",
    "team_key_passes_sum_3", "team_key_passes_sum_5",
    "team_influence_sum_3", "team_influence_sum_5",
    "team_bps_sum_3", "team_bps_sum_5",
]
for col in norm_cols:
    team_context[f"{col}_per_player"] = (
        team_context[col] / team_context["squad_size"].clip(lower=1)
    )

team_context_cols = (
    [c + "_per_player" for c in norm_cols] +
    [
        "teammates_scoring_3", "teammates_scoring_5",
        "teammates_assisting_3", "teammates_assisting_5",
        "teammates_playing_3", "teammates_playing_5",
    ]
)

assert team_context.duplicated(subset=["team", "date"]).sum() == 0, \
    "Duplicate (team, date) rows in team_context"

# Merge team context onto player level
df = df.merge(
    team_context[["team", "date"] + team_context_cols],
    on     = ["team", "date"],
    how    = "left",
    suffixes = ("", "_drop"),
)
df = df.drop(columns=[c for c in df.columns if c.endswith("_drop")])
df = df.loc[:, ~df.columns.duplicated()].reset_index(drop=True)

print(f"After team context merge — duplicate indices: {df.index.duplicated().sum()}")
print(f"Team context features added: {len(team_context_cols)}")

# =============================================================================
# 5. FEATURE SET
# =============================================================================
features = [
    # Context
    "is_home", "price", "gameweek", "own_league_position", "opp_league_position",

    # Minutes
    "form_minutes_3", "form_minutes_5", "season_minutes",

    # Attacking
    "form_goals_3", "form_goals_5",
    "form_xg_3", "form_xg_5",
    "form_shots_3", "form_shots_5",

    # Creativity
    "form_assists_3", "form_assists_5",
    "form_xa_3", "form_xa_5",
    "form_key_passes_y_3", "form_key_passes_y_5",

    # Bonus proxies
    "form_bps_3", "form_bps_5",
    "form_influence_3", "form_influence_5",
    "form_creativity_3", "form_creativity_5",

    # Opponent strength
    "opp_xg_conceded_form_3", "opp_xg_conceded_form_5",
    "opp_goals_conceded_form_3", "opp_goals_conceded_form_5",

    # Season totals
    "season_goals", "season_assists", "clean_sheets_season",

    # Previous season summary
    "prev_season_pts", "prev_season_goals", "prev_season_assists",
    "prev_season_minutes", "prev_season_xg", "prev_season_xa",
    "prev_season_games", "prev_season_complete",

    # ── NEW: Team-level inter-player features ─────────────────────────
    # Attacking context (per player normalised)
    "team_xg_sum_3_per_player", "team_xg_sum_5_per_player",
    "team_shots_sum_3_per_player", "team_shots_sum_5_per_player",
    "team_goals_sum_3_per_player", "team_goals_sum_5_per_player",

    # Creativity context
    "team_xa_sum_3_per_player", "team_xa_sum_5_per_player",
    "team_creativity_sum_3_per_player", "team_creativity_sum_5_per_player",
    "team_key_passes_sum_3_per_player", "team_key_passes_sum_5_per_player",

    # Influence context
    "team_influence_sum_3_per_player", "team_influence_sum_5_per_player",
    "team_bps_sum_3_per_player", "team_bps_sum_5_per_player",

    # In-form teammate counts
    "teammates_scoring_3", "teammates_scoring_5",
    "teammates_assisting_3", "teammates_assisting_5",
    "teammates_playing_3", "teammates_playing_5",
]

missing = [f for f in features if f not in df.columns]
if missing:
    raise ValueError(f"Missing feature columns: {missing}")




train_df = df[
    (df["season_x"] == PREV_SEASON) |
    ((df["season_x"] == SEASON) & (df["gameweek"] < TARGET_GW))
].copy()
params = {
    "n_estimators": 300,
    "max_depth": 6,
    "learning_rate": 0.05,
    "subsample": 0.8,
    "colsample_bytree": 0.8,
    "random_state": 42,
    "verbosity": 0
}
model = XGBRegressor(**params)

model.fit(train_df[features].fillna(0), train_df["total_points"])






def select_team(data, mode="consistent"):
    df_sim = data.copy()

    if mode == "consistent":
        df_sim["score"] = df_sim["sim_mean"] / (1 + df_sim["sim_std"])
    else:
        df_sim["score"] = df_sim["sim_mean"] * (1 + df_sim["sim_std"])

    df_sim["value_metric"] = df_sim["score"] / 0.5 * df_sim["price"]

    selected = []
    budget = BUDGET
    pos_count = {k: 0 for k in constraints}
    team_count = {}

    for _, row in df_sim.sort_values("value_metric", ascending=False).iterrows():
        pos = row["position_mapped"]

        if pos_count[pos] >= constraints[pos]:
            continue
        if team_count.get(row["team"], 0) >= MAX_PER_TEAM:
            continue
        if budget < row["price"]:
            continue

        selected.append(row)
        budget -= row["price"]
        pos_count[pos] += 1
        team_count[row["team"]] = team_count.get(row["team"], 0) + 1

        if sum(pos_count.values()) == 11:
            break

    return pd.DataFrame(selected)

def write_team_sheet(wb, df_team, sheet_name):
    ws = wb.create_sheet(title=str(sheet_name))

    cols = [
        "player", "position_mapped", "team", "opponent", "is_home",
        "price", "sim_mean", "sim_std", "total_points"
    ]

    df_out = df_team[cols].copy()

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    ws.append([])
    ws.append(["Total Price", df_team["price"].sum()])
    ws.append(["Total Expected Points", df_team["sim_mean"].sum()])
    ws.append(["Total Actual Points", df_team["total_points"].fillna(0).sum()])

def run_full_backtest(start_gw=1, end_gw=38):

    wb = Workbook()
    wb.remove(wb.active)

    results = []

    summary_rows = []
    all_players = []

    def score(team):
        errors = team["sim_mean"] - team["total_points"].fillna(0)

        mae = errors.abs().mean()
        bias = errors.mean()

        pred_avg = team["sim_mean"].sum()
        act_avg = team["total_points"].fillna(0).sum()

        return pred_avg, act_avg, mae, bias

    base_df = df.copy()
    season_cons_abs = 0
    season_risk_abs = 0
    for gw in range(start_gw, end_gw + 1):

        train = base_df[
            (base_df["season_x"] == PREV_SEASON) |
            ((base_df["season_x"] == SEASON) & (base_df["gameweek"] < gw))
        ].copy()

        test_raw = base_df[
            (base_df["season_x"] == SEASON) &
            (base_df["gameweek"] == gw)
        ].copy()

        test = test_raw.sort_values(["player", "gameweek"]).groupby("player").tail(1).copy()

        if len(train) == 0 or len(test) == 0:
            continue

        model.fit(train[features].fillna(0), train["total_points"])

        test = test.copy()
        test["expected_points"] = model.predict(test[features].fillna(0))
        test["sim_mean"] = test["expected_points"]
        test["sim_std"] = test["std_points"]

        cons_team = select_team(test, "consistent")
        risk_team = select_team(test, "risky")

        cons_team["mode"] = "Consistent"
        risk_team["mode"] = "Risky"
        cons_team["GW"] = gw
        risk_team["GW"] = gw

        all_players.append(cons_team.copy())
        all_players.append(risk_team.copy())

        print("\nGW", gw)

        print("\nCONSISTENT TEAM")
        print(cons_team[[
            "player", "position_mapped", "team", "price", "sim_mean", "sim_std", "total_points"
        ]].to_string(index=False))

        print("\nRISKY TEAM")
        print(risk_team[[
            "player", "position_mapped", "team", "price", "sim_mean", "sim_std", "total_points"
        ]].to_string(index=False))

        cp, ca, cmae, cbias = score(cons_team)
        rp, ra, rmae, rbias = score(risk_team)

        season_cons_abs += abs(cp - ca)
        season_risk_abs += abs(rp - ra)

        results.append({
            "GW": gw,
            "cons_pred": cp,
            "cons_actual": ca,
            "cons_mae": cmae,
            "cons_bias": cbias,
            "risk_pred": rp,
            "risk_actual": ra,
            "risk_mae": rmae,
            "risk_bias": rbias
        })

        summary_rows.append({
            "GW": gw,
            "cons_points": cons_team["total_points"].fillna(0).sum(),
            "risk_points": risk_team["total_points"].fillna(0).sum(),
            "cons_expected_points": cons_team["sim_mean"].sum(),
            "risk_expected_points": risk_team["sim_mean"].sum(),
            "cons_mae": cmae,
            "risk_mae": rmae
        })

        write_team_sheet(wb, cons_team, f"GW{gw}_Consistent")
        write_team_sheet(wb, risk_team, f"GW{gw}_Risky")

    summary_df = pd.DataFrame(summary_rows)

    total_cons_points = summary_df["cons_points"].sum()
    total_risk_points = summary_df["risk_points"].sum()

    total_cons_expected = summary_df["cons_expected_points"].sum()
    total_risk_expected = summary_df["risk_expected_points"].sum()

    abs_error_cons = season_cons_abs / GWs
    abs_error_risk = season_risk_abs / GWs

    avg_cons_mae = summary_df["cons_mae"].mean()
    avg_risk_mae = summary_df["risk_mae"].mean()

    players_df = pd.concat(all_players)

    players_df["error"] = players_df["sim_mean"] - players_df["total_points"].fillna(0)
    players_df["abs_error"] = players_df["error"].abs()

    pos_summary = players_df.groupby(["mode", "position_mapped"]).agg({
        "total_points": "sum",
        "abs_error": "mean"
    }).reset_index()

    ws = wb.create_sheet(title="Summary")

    ws.append(["Metric", "Consistent", "Risky"])
    ws.append(["Total Points", total_cons_points, total_risk_points])
    ws.append(["Predicted Points", total_cons_expected, total_risk_expected])
    ws.append(["Difference", abs_error_cons, abs_error_risk])
    ws.append(["Average MAE", avg_cons_mae, avg_risk_mae])

    ws.append([])
    ws.append(["Position Breakdown"])
    ws.append(["Mode", "Position", "Total Points", "Avg MAE"])

    for _, row in pos_summary.iterrows():
        ws.append([
            row["mode"],
            row["position_mapped"],
            row["total_points"],
            row["abs_error"]
        ])
    summary_ws = wb["Summary"]
    wb._sheets.remove(summary_ws)
    wb._sheets.insert(0, summary_ws)

    wb.save("FPL_All_Gameweek_Teams.xlsx")

    return pd.DataFrame(results)

backtest_df = run_full_backtest()

plot_df = backtest_df[
    (backtest_df["GW"] >= GRAPH_START_GW) &
    (backtest_df["GW"] <= GRAPH_END_GW)
].copy()

plt.figure()
plt.plot(plot_df["GW"], plot_df["cons_pred"], label="Cons Pred")
plt.plot(plot_df["GW"], plot_df["cons_actual"], label="Cons Actual")
plt.plot(plot_df["GW"], plot_df["risk_pred"], linestyle="--", label="Risk Pred")
plt.plot(plot_df["GW"], plot_df["risk_actual"], linestyle="--", label="Risk Actual")
plt.legend()
plt.show()

plt.figure()
plt.plot(plot_df["GW"], plot_df["cons_mae"], label="Cons MAE")
plt.plot(plot_df["GW"], plot_df["risk_mae"], label="Risk MAE")
plt.legend()
plt.show()

plt.figure()
plot_df["cons_mae"] = plot_df["cons_mae"].rolling(ROLLING_WINDOW).mean()
plot_df["risk_mae"] = plot_df["risk_mae"].rolling(ROLLING_WINDOW).mean()
plt.plot(plot_df["GW"], plot_df["cons_mae"], label="Cons MAE Smoothed")
plt.plot(plot_df["GW"], plot_df["risk_mae"], label="Risk MAE Smoothed")
plt.legend()
plt.show()

plt.figure()
plot_df["cons_bias"] = plot_df["cons_bias"].rolling(ROLLING_WINDOW).mean()
plot_df["risk_bias"] = plot_df["risk_bias"].rolling(ROLLING_WINDOW).mean()
plt.plot(plot_df["GW"], plot_df["cons_bias"], label="Cons Bias")
plt.plot(plot_df["GW"], plot_df["risk_bias"], label="Risk Bias")
plt.axhline(0)
plt.legend()
plt.show()